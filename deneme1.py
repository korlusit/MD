import sys
from PyQt6.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget
import openpyxl

class ExcelViewer(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Mav Excel Görüntüleyici")
        self.setGeometry(100, 100, 800, 600)

        self.tableWidget = QTableWidget()
        self.setCentralWidget(self.tableWidget)

    def load_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        self.tableWidget.setRowCount(sheet.max_row)
        self.tableWidget.setColumnCount(sheet.max_column)

        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                self.tableWidget.setItem(row - 1, col - 1, QTableWidgetItem(str(cell_value)))

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelViewer()
    window.show()

    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        window.load_excel(file_path)

    sys.exit(app.exec_())
