import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Side, Border
import os
import sqlite3
import datetime

def excel():
    try:
        # Masaüstü yolunu al
        masaustu = os.path.join(os.path.expanduser("~"), "Desktop")
        excel_klasor = os.path.join(masaustu, "Excel")

        # Excel klasörü yoksa oluştur
        if not os.path.exists(excel_klasor):
            os.makedirs(excel_klasor)

        # SQLite veritabanına bağlan
        baglan = sqlite3.connect("database.db")
        c = baglan.cursor()
        data = c.execute("SELECT * FROM kisiler")

        # CSV dosyasına yaz
        with open('output.csv', 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Ad", "Tarih", "GSM", "Bolum", "Ek Not", "Unvan", "Ders", "ID"])
            writer.writerows(data)

        # CSV'den pandas DataFrame'e oku
        df = pd.read_csv('output.csv', encoding='utf-8')

        # Excel dosyası oluştur
        wb = Workbook()
        ws = wb.active

        # Başlık ve satır stilleri
        header_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        row_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Başlık yaz
        for col_idx, header in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.border = border

        # Verileri yaz
        for row_idx, row in enumerate(df.itertuples(), start=2):
            for col_idx, value in enumerate(row[1:], start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = border
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

        # Dosya adını hazırla
        dosya_adi = os.path.join(excel_klasor, "Mba_{}_Kisiler.xlsx".format(
            datetime.datetime.now().strftime("%d-%m-%Y_%H-%M-%S-%f")
        ))

        # Excel dosyasını kaydet
        wb.save(dosya_adi)

        # Geçici CSV dosyasını sil
        os.remove('output.csv')

        print(f"✅ Excel dosyası başarıyla kaydedildi: {dosya_adi}")

    except Exception as e:
        print(f"❌ Bir hata oluştu: {e}")
